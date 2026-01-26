import os
import re
import glob
from copy import copy as copy_style
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment


# ------------------------------------------------------------
# Ordner / Layouts
# ------------------------------------------------------------

LAYOUT_DIR = "Layouts"
OUTPUT_DIR = "Ausgabedateien"

INTERNAL_HEADER_TEXT = "NUR FÜR DEN INTERNEN DIENSTGEBRAUCH"

RAW_SHEET_NAMES = {
    1: "XML-Tab1-Land",
    2: "XML-Tab2-Land",
    3: "XML-Tab3-Land",
    5: "XML-Tab5-Land",
}

# Layout-Dateien (liegen in ./Layouts/)
TEMPLATES = {
    1: {"ext": "Tabelle-1-Layout_g.xlsx", "int": "Tabelle-1-Layout_INTERN.xlsx"},
    2: {"ext": "Tabelle-2-Layout_g.xlsx", "int": "Tabelle-2-Layout_INTERN.xlsx"},
    3: {"ext": "Tabelle-3-Layout_g.xlsx", "int": "Tabelle-3-Layout_INTERN.xlsx"},
    5: {"ext": "Tabelle-5-Layout_g.xlsx", "int": "Tabelle-5-Layout_INTERN.xlsx"},
}


# ------------------------------------------------------------
# Helper
# ------------------------------------------------------------

def is_numeric_like(v):
    """Erkennt Zahlen/Platzhalter in Zellen (inkl. '-', 'X', mit Punkt/Komma)."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip()
        if s in ("-", "X", ""):
            return True
        s2 = s.replace(".", "").replace(" ", "").replace(",", "")
        return s2.isdigit()
    return False


def extract_month_from_raw(ws, table_no):
    """
    Holt Monats-/Zeitraum-Text aus Rohblatt.
    (Diese Positionen waren bei uns konsistent)
    """
    if table_no == 1:
        return ws.cell(row=3, column=1).value
    if table_no in (2, 3):
        return ws.cell(row=4, column=1).value
    if table_no == 5:
        return ws.cell(row=3, column=1).value
    return None


def extract_stand_from_raw(ws, max_search_rows=60):
    """Sucht 'Stand:' in den letzten Zeilen des Rohblatts."""
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(max_row, max(max_row - max_search_rows, 1) - 1, -1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "Stand:" in v:
                return v.strip()
    return None


def get_merged_secondary_checker(ws):
    """True, wenn Zelle in Merge liegt aber NICHT die Top-Left-Zelle ist."""
    merged = list(ws.merged_cells.ranges)

    def is_secondary(row, col):
        for rg in merged:
            if rg.min_row <= row <= rg.max_row and rg.min_col <= col <= rg.max_col:
                return not (row == rg.min_row and col == rg.min_col)
        return False

    return is_secondary


def update_footer_with_stand_and_copyright(ws, stand_text):
    """
    Aktualisiert:
      - Copyright-Jahr -> aktuelles Jahr
      - Stand-Text -> in die Stand-Spalte der Copyright-Zeile (oder letzte Spalte)
      - entfernt doppelte Stand-Zeilen (falls irgendwo sonst im Blatt)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    cy = datetime.now().year

    # Copyright-Zeile finden (Spalte A)
    copyright_row = None
    for r in range(max_row, 0, -1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and ("(C)opyright" in v or "Copyright" in v):
            # Jahr ersetzen/setzen
            txt = v
            txt = re.sub(r"\(C\)opyright\s+\d{4}", f"(C)opyright {cy}", txt)
            if "(C)opyright" not in txt:
                txt = f"(C)opyright {cy} Bayerisches Landesamt für Statistik"
            ws.cell(row=r, column=1).value = txt
            copyright_row = r
            break

    if copyright_row is None:
        # keine Copyrightzeile: am Ende Leerzeile + Zeile einfügen
        ws.append([])
        ws.append([])
        copyright_row = ws.max_row
        ws.cell(row=copyright_row, column=1).value = f"(C)opyright {cy} Bayerisches Landesamt für Statistik"

    # andere Stand:-Vorkommen entfernen
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().startswith("Stand:") and r != copyright_row:
                ws.cell(row=r, column=c).value = ""

    if not stand_text:
        return

    # Stand-Spalte suchen (in Copyright-Zeile), sonst letzte Spalte
    stand_col = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=copyright_row, column=c).value
        if isinstance(v, str) and "Stand:" in v:
            stand_col = c
            break
    if stand_col is None:
        stand_col = max_col

    # Stil von Copyright-Zelle übernehmen
    cop_cell = ws.cell(row=copyright_row, column=1)
    tgt = ws.cell(row=copyright_row, column=stand_col)
    tgt.value = stand_text.strip()

    tgt.font = copy_style(cop_cell.font)
    tgt.border = copy_style(cop_cell.border)
    tgt.fill = copy_style(cop_cell.fill)
    tgt.number_format = cop_cell.number_format
    tgt.protection = copy_style(cop_cell.protection)
    tgt.alignment = Alignment(horizontal="right",
                             vertical=cop_cell.alignment.vertical if cop_cell.alignment else "center")

    # falls doppelt "Stand:" im Text
    if isinstance(tgt.value, str) and tgt.value.count("Stand:") > 1:
        i = tgt.value.find("Stand:")
        tgt.value = "Stand:" + tgt.value[i + len("Stand:"):].strip()


def out_path_for(raw_path, suffix):
    """Schreibt in OUTPUT_DIR"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    base = os.path.basename(raw_path).replace(".xlsx", f"{suffix}.xlsx")
    return os.path.join(OUTPUT_DIR, base)


def load_layout(path_in_layouts):
    p = os.path.join(LAYOUT_DIR, path_in_layouts)
    if not os.path.exists(p):
        raise FileNotFoundError(f"Layout-Datei fehlt: {p}")
    return openpyxl.load_workbook(p)


# ------------------------------------------------------------
# Tabelle 1
# ------------------------------------------------------------

def process_table1(raw_path, tmpl_ext, tmpl_int):
    print(f"Verarbeite Tabelle 1 aus '{os.path.basename(raw_path)}' ...")

    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[1]]

    month_text = extract_month_from_raw(ws_raw, 1)
    stand_text = extract_stand_from_raw(ws_raw)

    # Daten- und Fußnotenbereich bestimmen
    def detect_data_and_footer(ws, numeric_col=4):
        max_row = ws.max_row
        first_data = None
        for r in range(1, max_row + 1):
            if is_numeric_like(ws.cell(row=r, column=numeric_col).value):
                first_data = r
                break
        if first_data is None:
            first_data = 1

        footnote_start = max_row + 1
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().startswith("-"):
                footnote_start = r
                break
        return first_data, footnote_start

    fdr_raw, ft_raw = detect_data_and_footer(ws_raw, numeric_col=4)

    # -------- EXTERN --------
    wb_ext = load_layout(tmpl_ext)
    ws_ext = wb_ext[wb_ext.sheetnames[0]]
    ws_ext.cell(row=3, column=1).value = month_text

    is_sec = get_merged_secondary_checker(ws_ext)
    fdr_ext, ft_ext = detect_data_and_footer(ws_ext, numeric_col=4)

    n_rows = min(ft_raw - fdr_raw, ft_ext - fdr_ext)
    for off in range(n_rows):
        r_raw = fdr_raw + off
        r_ext = fdr_ext + off
        for c in range(1, ws_ext.max_column + 1):
            if is_sec(r_ext, c):
                continue
            ws_ext.cell(row=r_ext, column=c).value = ws_raw.cell(row=r_raw, column=c).value

    update_footer_with_stand_and_copyright(ws_ext, stand_text)

    out_ext = out_path_for(raw_path, "_g")
    wb_ext.save(out_ext)
    print(f"  -> Extern: {out_ext}")

    # -------- INTERN --------
    wb_int = load_layout(tmpl_int)
    ws_int = wb_int[wb_int.sheetnames[0]]
    ws_int.cell(row=1, column=1).value = INTERNAL_HEADER_TEXT
    ws_int.cell(row=5, column=1).value = month_text

    is_sec_i = get_merged_secondary_checker(ws_int)
    fdr_int, ft_int = detect_data_and_footer(ws_int, numeric_col=4)

    n_rows = min(ft_raw - fdr_raw, ft_int - fdr_int)
    for off in range(n_rows):
        r_raw = fdr_raw + off
        r_int = fdr_int + off
        for c in range(1, ws_int.max_column + 1):
            if is_sec_i(r_int, c):
                continue
            ws_int.cell(row=r_int, column=c).value = ws_raw.cell(row=r_raw, column=c).value

    update_footer_with_stand_and_copyright(ws_int, stand_text)

    out_int = out_path_for(raw_path, "_INTERN")
    wb_int.save(out_int)
    print(f"  -> Intern: {out_int}")


# ------------------------------------------------------------
# Tabelle 2 & 3  (WICHTIG: FIX = ab Spalte B kopieren!)
# ------------------------------------------------------------

def process_table2_or_3(table_no, raw_path, tmpl_ext, tmpl_int):
    print(f"Verarbeite Tabelle {table_no} aus '{os.path.basename(raw_path)}' ...")

    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[table_no]]

    month_text = extract_month_from_raw(ws_raw, table_no)
    stand_text = extract_stand_from_raw(ws_raw)

    # FIX: Startspalte = 2 (Spalte B)  -> damit Spalte B nicht aus Layout bleibt
    START_COL = 2

    def detect_data_and_footer(ws, numeric_col):
        max_row = ws.max_row
        first_data = None
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=numeric_col).value
            if is_numeric_like(v):
                first_data = r
                break
        if first_data is None:
            first_data = 1

        footnote_start = max_row + 1
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().startswith("-"):
                footnote_start = r
                break
        return first_data, footnote_start

    fdr_raw, ft_raw = detect_data_and_footer(ws_raw, numeric_col=START_COL)

    def fill_numeric(ws_t):
        is_sec = get_merged_secondary_checker(ws_t)
        fdr_t, ft_t = detect_data_and_footer(ws_t, numeric_col=START_COL)

        n_rows = min(ft_raw - fdr_raw, ft_t - fdr_t)
        for off in range(n_rows):
            r_raw = fdr_raw + off
            r_t = fdr_t + off

            # FIX: ab Spalte B (START_COL) überschreiben
            for c in range(START_COL, ws_t.max_column + 1):
                if is_sec(r_t, c):
                    continue
                ws_t.cell(row=r_t, column=c).value = ws_raw.cell(row=r_raw, column=c).value

    # -------- EXTERN --------
    wb_ext = load_layout(tmpl_ext)
    ws_ext = wb_ext[wb_ext.sheetnames[0]]
    ws_ext.cell(row=3, column=1).value = month_text
    fill_numeric(ws_ext)
    update_footer_with_stand_and_copyright(ws_ext, stand_text)

    out_ext = out_path_for(raw_path, "_g")
    wb_ext.save(out_ext)
    print(f"  -> Extern: {out_ext}")

    # -------- INTERN --------
    wb_int = load_layout(tmpl_int)
    ws_int = wb_int[wb_int.sheetnames[0]]
    ws_int.cell(row=1, column=1).value = INTERNAL_HEADER_TEXT
    ws_int.cell(row=6, column=1).value = month_text
    fill_numeric(ws_int)
    update_footer_with_stand_and_copyright(ws_int, stand_text)

    out_int = out_path_for(raw_path, "_INTERN")
    wb_int.save(out_int)
    print(f"  -> Intern: {out_int}")


# ------------------------------------------------------------
# Tabelle 5 (5 Blätter / 5 Blöcke)
# ------------------------------------------------------------

def process_table5(raw_path, tmpl_ext, tmpl_int):
    print(f"Verarbeite Tabelle 5 aus '{os.path.basename(raw_path)}' ...")

    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[5]]

    month_text = extract_month_from_raw(ws_raw, 5)
    stand_text = extract_stand_from_raw(ws_raw)

    max_row = ws_raw.max_row

    # Blockstart via "Bayern x)" in Spalte B
    starts = []
    for r in range(1, max_row + 1):
        v = ws_raw.cell(row=r, column=2).value
        if isinstance(v, str) and re.match(r"Bayern\s+\d\)", v.strip()):
            starts.append(r)

    block_ranges = []
    for i, start in enumerate(starts):
        end = (starts[i + 1] - 1) if i < len(starts) - 1 else max_row
        last_nonempty = start
        for r in range(start, end + 1):
            if any(ws_raw.cell(row=r, column=c).value not in (None, "") for c in range(1, 12)):
                last_nonempty = r
        block_ranges.append((start, last_nonempty))

    # kopiert nur Zahlen C..H
    def fill_sheet_from_block(ws_t, start_row, end_row):
        is_sec = get_merged_secondary_checker(ws_t)

        # erste Datenzeile im Template: erste Zeile mit numerischem Wert in Spalte 3
        first_data_t = None
        for r in range(1, ws_t.max_row + 1):
            if is_numeric_like(ws_t.cell(row=r, column=3).value):
                first_data_t = r
                break
        if first_data_t is None:
            return

        raw_r = start_row
        t_r = first_data_t
        while raw_r <= end_row and t_r <= ws_t.max_row:
            for c in range(3, 9):  # C..H
                if is_sec(t_r, c):
                    continue
                ws_t.cell(row=t_r, column=c).value = ws_raw.cell(row=raw_r, column=c).value
            raw_r += 1
            t_r += 1

    # -------- EXTERN --------
    wb_ext = load_layout(tmpl_ext)
    for i, (start, end) in enumerate(block_ranges):
        if i >= len(wb_ext.worksheets):
            break
        ws_t = wb_ext.worksheets[i]
        ws_t.cell(row=3, column=1).value = month_text
        fill_sheet_from_block(ws_t, start, end)
        update_footer_with_stand_and_copyright(ws_t, stand_text)

    out_ext = out_path_for(raw_path, "_g")
    wb_ext.save(out_ext)
    print(f"  -> Extern: {out_ext}")

    # -------- INTERN --------
    wb_int = load_layout(tmpl_int)
    for i, (start, end) in enumerate(block_ranges):
        if i >= len(wb_int.worksheets):
            break
        ws_t = wb_int.worksheets[i]
        ws_t.cell(row=1, column=1).value = INTERNAL_HEADER_TEXT
        ws_t.cell(row=5, column=1).value = month_text
        fill_sheet_from_block(ws_t, start, end)
        update_footer_with_stand_and_copyright(ws_t, stand_text)

    out_int = out_path_for(raw_path, "_INTERN")
    wb_int.save(out_int)
    print(f"  -> Intern: {out_int}")


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main():
    print("Starte Tabellen-Formatter (1,2,3,5 – INTERN & EXTERN)...")
    print(f"Arbeitsverzeichnis: {os.getcwd()}")
    print()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for table_no in (1, 2, 3, 5):
        pattern = f"Tabelle-{table_no}-Land_*.xlsx"
        candidates = sorted(glob.glob(pattern))

        # nur Rohdateien
        raw_files = [
            f for f in candidates
            if not f.endswith("_g.xlsx") and not f.endswith("_INTERN.xlsx")
        ]

        if not raw_files:
            print(f"Keine Rohdateien für Tabelle {table_no} gefunden ({pattern}) – übersprungen.\n")
            continue

        tmpl = TEMPLATES[table_no]
        tmpl_ext = tmpl["ext"]
        tmpl_int = tmpl["int"]

        for raw_path in raw_files:
            if table_no == 1:
                process_table1(raw_path, tmpl_ext, tmpl_int)
            elif table_no in (2, 3):
                process_table2_or_3(table_no, raw_path, tmpl_ext, tmpl_int)
            elif table_no == 5:
                process_table5(raw_path, tmpl_ext, tmpl_int)

            print()

    print("Fertig. 🎉")


if __name__ == "__main__":
    main()
